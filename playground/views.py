from django.shortcuts import render
from django.http import HttpResponse

import os
from io import BytesIO


from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


from .models import Keys, MiscItem


def say_hello(request):
    "pull data, transform, send mails etc"
    return render(request, 'hello.html')

from django.shortcuts import render

def export_dashboard(request):
    return render(request, 'export_dashboard.html')


#change


from openpyxl import Workbook

from django.shortcuts import render
from .models import Keys  # make sure the import is correct

def keys_list(request):
    keys = Keys.objects.all()
    return render(request, 'keys_list.html', {'keys': keys})

from .models import MiscItem

def misc_list(request):
    items = MiscItem.objects.all()
    return render(request, 'misc_list.html', {'items': items})

from .models import BlockCStationaryItem

def stationary_list(request):
    items = BlockCStationaryItem.objects.all()
    return render(request, 'stationary_list.html', {'items': items})

from .models import BlockAConsumableItem
def consumable_list(request):
    items = BlockAConsumableItem.object.all()
    return render(request, 'consumable_list.html',  {'items': items})

from .models import FixedAssets

def asset_list(request):
    items = FixedAssets.objects.all()
    return render(request, 'asset_list.html', {'items': items})


from .models import PPEs

def ppe_list(request):
    items = PPEs.objects.all()
    return render(request, 'ppe_list.html', {'items': items})


from .models import SpareItems

def spare_list(request):
    items = SpareItems.objects.all()
    return render(request, 'spare_list.html', {'items': items})

from .models import TeaItems

def tea_list(request):
    items = TeaItems.objects.all()
    return render(request, 'tea_list.html', {'items': items})


###########################################################################


def export_keys_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Keys Data"

    # Query your DB model
    qs = Keys.objects.all()

    filename = "keys_export.xlsx"

    # headers
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 'quantity_mentioned',
        'actual_quantity', 'location', 'person', 'dispatch_loc', 'image', 'remarks'
    ]
    ws.append(headers)

    # column width
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)  # image column index (1-based)

    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.quantity_mentioned,
            obj.actual_quantity,
            obj.location,
            obj.person,
            obj.dispatch_loc,
            '',  # placeholder for image
            obj.remarks,
        ]
        ws.append(row)

        # Insert image if present
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80
                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell
                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60
                except Exception as e:
                    print('Failed to add image for', obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response



from playground.models import MiscItem   # make sure this is correct


def export_miscitem_excel(request):

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Misc Items"

    # Query the database
    qs = MiscItem.objects.all()

    # File name for download
    filename = "misc_items_export.xlsx"

    # Headers
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
        'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks'
    ]
    ws.append(headers)

    # Set column width
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)  # column number of the image field

    # Add rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.amount_per_box_pack_roll,
            obj.full_boxes_packs_rolls,
            obj.open,
            obj.in_stock,
            obj.location,
            '',  # placeholder for image
            obj.remarks,
        ]
        ws.append(row)

        # Add image (if exists)
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path

            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80

                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell

                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60

                except Exception as e:
                    print("Failed to add image for:", obj, e)

        row_index += 1

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Prepare response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response


from playground.models import FixedAssets   # change if your model is in another app


def export_fixedassets_excel(request):

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fixed Assets"

    # Query the database
    qs = FixedAssets.objects.all()

    # File name
    filename = "fixed_assets_export.xlsx"

    # Headers (same as MiscItem)
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
        'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks'
    ]
    ws.append(headers)

    # Column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)  # image column index

    # Add data rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.amount_per_box_pack_roll,
            obj.full_boxes_packs_rolls,
            obj.open,
            obj.in_stock,
            obj.location,
            '',  # placeholder for image
            obj.remarks,
        ]
        ws.append(row)

        # Insert image (if exists)
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path

            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80

                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell

                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60

                except Exception as e:
                    print("Failed to add image:", obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return file response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response




from playground.models import PPEs     # adjust import if model is in another app


def export_ppes_excel(request):

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PPEs"

    # Query database
    qs = PPEs.objects.all()

    # File name
    filename = "ppes_export.xlsx"

    # Same headers
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
        'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks'
    ]
    ws.append(headers)

    # Set column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)

    # Add data rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.amount_per_box_pack_roll,
            obj.full_boxes_packs_rolls,
            obj.open,
            obj.in_stock,
            obj.location,
            '',  # placeholder for image column
            obj.remarks,
        ]
        ws.append(row)

        # Insert image (if exists)
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path

            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80

                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell

                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60

                except Exception as e:
                    print("Failed to add image:", obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return file
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response




from playground.models import SpareItems   # adjust import if needed


def export_spareitems_excel(request):

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Spare Items"

    # Query database
    qs = SpareItems.objects.all()

    # File name
    filename = "spare_items_export.xlsx"

    # Headers (same as others)
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
        'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks'
    ]
    ws.append(headers)

    # Set column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)

    # Add rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.amount_per_box_pack_roll,
            obj.full_boxes_packs_rolls,
            obj.open,
            obj.in_stock,
            obj.location,
            '',  # image placeholder
            obj.remarks,
        ]
        ws.append(row)

        # Add image if exists
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80

                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell

                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60

                except Exception as e:
                    print("Failed to add image:", obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response



from playground.models import Keys  # adjust app name if different


def export_keys_excel(request):
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Keys"

    # Query the database
    qs = Keys.objects.all()

    # File name for download
    filename = "keys_export.xlsx"

    # Headers matching your Keys model fields
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom',
        'quantity_mentioned', 'actual_quantity', 'location',
        'person', 'dispatch_loc', 'image', 'remarks'
    ]
    ws.append(headers)

    # Set column widths for readability
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)  # image column index (1-based)

    # Add data rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.quantity_mentioned,
            obj.actual_quantity,
            obj.location,
            obj.person,
            obj.dispatch_loc,
            '',  # image placeholder
            obj.remarks,
        ]
        ws.append(row)

        # Add image if exists
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80
                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell
                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60
                except Exception as e:
                    print("Failed to add image for:", obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response




from playground.models import TeaItems   # adjust app name if different


def export_teaitems_excel(request):
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Tea Items"

    # Query the database
    qs = TeaItems.objects.all()

    # File name
    filename = "tea_items_export.xlsx"

    # Headers matching your TeaItems model fields
    headers = [
        'sr_no', 'item_name', 'item_description', 'uom', 
        'amount_per_box_pack_roll', 'full_boxes_packs_rolls', 'open',
        'old_stock', 'in_stock', 'location', 'person', 'dispatch_loc',
        'image', 'remarks'
    ]
    ws.append(headers)

    # Set column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)  # image column index

    # Add data rows
    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            obj.amount_per_box_pack_roll,
            obj.full_boxes_packs_rolls,
            obj.open,
            obj.old_stock,
            obj.in_stock,
            obj.location,
            obj.person,
            obj.dispatch_loc,
            '',  # image placeholder
            obj.remarks,
        ]
        ws.append(row)

        # Insert image if exists
        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80
                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell
                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60
                except Exception as e:
                    print("Failed to add image for:", obj, e)

        row_index += 1

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response


# ----------------- Consumables -----------------
def export_consumables_excel(request):
    qs = Consumables.objects.all()
    headers = ['sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
               'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks']

    wb = Workbook()
    ws = wb.active
    ws.title = "Consumables"
    ws.append(headers)

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)

    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            getattr(obj, 'amount_per_box_pack_roll', ''),
            getattr(obj, 'full_boxes_packs_rolls', ''),
            getattr(obj, 'open', ''),
            getattr(obj, 'in_stock', ''),
            obj.location,
            '',  # image placeholder
            obj.remarks,
        ]
        ws.append(row)

        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80
                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell
                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60
                except Exception as e:
                    print('Failed to add image for', obj, e)
        row_index += 1

    filename = "Consumables.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# ----------------- Stationary -----------------
def export_stationary_excel(request):
    qs = Stationary.objects.all()
    headers = ['sr_no', 'item_name', 'item_description', 'uom', 'amount_per_box_pack_roll',
               'full_boxes_packs_rolls', 'open', 'in_stock', 'location', 'image', 'remarks']

    wb = Workbook()
    ws = wb.active
    ws.title = "Stationary"
    ws.append(headers)

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    row_index = 2
    IMAGE_COL = len(headers)

    for obj in qs:
        row = [
            obj.sr_no,
            obj.item_name,
            (obj.item_description or '')[:32767],
            obj.uom,
            getattr(obj, 'amount_per_box_pack_roll', ''),
            getattr(obj, 'full_boxes_packs_rolls', ''),
            getattr(obj, 'open', ''),
            getattr(obj, 'in_stock', ''),
            obj.location,
            '',  # image placeholder
            obj.remarks,
        ]
        ws.append(row)

        if getattr(obj, 'image') and obj.image:
            image_path = obj.image.path
            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 120
                    img.height = 80
                    anchor_cell = f"{get_column_letter(IMAGE_COL)}{row_index}"
                    img.anchor = anchor_cell
                    ws.add_image(img)
                    ws.row_dimensions[row_index].height = 60
                except Exception as e:
                    print('Failed to add image for', obj, e)
        row_index += 1

    filename = "Stationary.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

